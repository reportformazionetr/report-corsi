import pandas as pd
import os
from typing import List
from models.course import Course
from utils.logger import logger
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

class ReportExporter:
    """Exports populated Course lists to a highly-polished, styled Excel spreadsheet."""
    
    @staticmethod
    def to_dataframe(courses: List[Course]) -> pd.DataFrame:
        """Converts Course models to a pandas DataFrame with matching specification headers."""
        data = []
        for c in courses:
            data.append({
                "N.": c.n,
                "Ente proponente": c.ente_proponente,
                "Area": c.area,
                "Ambito": c.ambito,
                "Titolo": c.titolo,
                "Edizioni concluse": c.edizioni_concluse,
                "Edizioni da completare": c.edizioni_da_completare,
                "Fonte di finanziamento": c.fonte_finanziamento,
                "Crediti formativi": c.crediti_formativi if c.crediti_formativi > 0 else None,
                "Ore svolte": c.ore_svolte,
                "Partecipanti effettivi": c.partecipanti_effettivi,
                "Numero partecipanti ECM": c.numero_partecipanti,
                "Spesa sostenuta": c.spesa_sostenuta,
                "Dettaglio Calcolo": c.dettaglio_calcolo
            })
        return pd.DataFrame(data)

    @staticmethod
    def export(courses: List[Course], file_path: str):
        """Generates a styled Excel sheet at file_path."""
        logger.info(f"Avvio esportazione di {len(courses)} corsi in Excel...")
        
        df = ReportExporter.to_dataframe(courses)
        
        # Write to excel using pandas first
        df.to_excel(file_path, index=False, sheet_name="Report Annuale")
        
        # Load workbook with openpyxl to apply custom formatting
        wb = openpyxl.load_workbook(file_path)
        ws = wb["Report Annuale"]
        
        # Styles definition
        font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        fill_header = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid") # Deep navy header
        
        font_data = Font(name="Calibri", size=11)
        align_left = Alignment(horizontal="left", vertical="center")
        align_right = Alignment(horizontal="right", vertical="center")
        
        border_side = Side(border_style="thin", color="D9D9D9")
        cell_border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
        
        # Format columns: headers
        ws.row_dimensions[1].height = 28 # Taller header row
        for col_idx in range(1, len(df.columns) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = cell_border
            
        # Format data cells
        for r_idx in range(2, len(df) + 2):
            ws.row_dimensions[r_idx].height = 20 # Comfortable row heights
            
            for col_idx, col_name in enumerate(df.columns, start=1):
                cell = ws.cell(row=r_idx, column=col_idx)
                cell.font = font_data
                cell.border = cell_border
                
                # Check column type to set alignments and number formats
                # Spesa sostenuta (Col 12)
                if col_name == "Spesa sostenuta":
                    cell.alignment = align_right
                    cell.number_format = "$#,##0.00" # Or just #,##0.00 €
                    # Let's use currency-independent standard format with 2 decimals
                    cell.number_format = "#,##0.00"
                # Crediti formativi, Ore svolte (Col 9, 10)
                elif col_name in ["Crediti formativi", "Ore svolte"]:
                    cell.alignment = align_right
                    cell.number_format = "#,##0.0"
                # N., Edizioni concluse, Numero partecipanti ECM, Partecipanti effettivi
                elif col_name in ["N.", "Edizioni concluse", "Numero partecipanti ECM", "Partecipanti effettivi"]:
                    cell.alignment = align_right
                    cell.number_format = "0"
                # Texts
                else:
                    cell.alignment = align_left
                    
        # Auto-fit columns width dynamically
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            
            for cell in col:
                val = str(cell.value or '')
                # If cell has newline, count the longest line
                lines = val.split('\n')
                for line in lines:
                    if len(line) > max_len:
                        max_len = len(line)
                        
            # Set width with extra padding
            ws.column_dimensions[col_letter].width = max(max_len + 4, 12)
            
        # Save modifications
        wb.save(file_path)
        
        # Log completion safely checking if it's a file path or a buffer
        if isinstance(file_path, str):
            filename = os.path.basename(file_path)
        else:
            filename = "buffer di memoria"
        logger.info(f"Esportazione Excel completata con successo in: {filename}")
